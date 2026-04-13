#!/usr/bin/env python3
"""
import_suppliers.py  <file_path>

Lê um arquivo CSV (separador ;) ou XLSX e retorna as linhas em JSON para stdout.
Segue o template de colunas do sistema:
  Nome;Razão Social;CNPJ;Email;Descrição;Status;Código;Inscrição Estadual;
  Email principal;Telefone principal;Data fundação/Aniversário;CEP;Estado;Cidade;
  Endereço;Número;Bairro;Complemento;Nome Contato;E-mail Contato

Saída:
  { "rows": [...], "count": N }   sucesso
  { "error": "mensagem" }         erro
"""
import sys
import json
import os
import re


# ──────────────────────────────────────────────────────────────────────────────
# Normaliza cabeçalho → nome do campo interno
# Aceita tanto o template novo quanto cabeçalhos já normalizados (re-importação)
# ──────────────────────────────────────────────────────────────────────────────
HEADER_MAP = {
    # Template original
    'nome':                        'supplier_name',
    'razão social':                'supplier_legal_name',
    'razao social':                'supplier_legal_name',
    'cnpj':                        'supplier_legal_identifier',
    'email':                       'email_fallback',        # usado só se email principal vazio
    'descrição':                   'description',
    'descricao':                   'description',
    'status':                      '_status',               # ignorado
    'código':                      '_codigo',               # usado como id
    'codigo':                      '_codigo',
    'inscrição estadual':          '_ie',                   # ignorado
    'inscricao estadual':          '_ie',
    'email principal':             'supplier_email',
    'e-mail principal':            'supplier_email',
    'telefone principal':          'supplier_telephone',
    'data fundação/aniversário':   '_data',                 # ignorado
    'data fundacao/aniversario':   '_data',
    'cep':                         '_cep',
    'estado':                      '_estado',
    'cidade':                      '_cidade',
    'endereço':                    '_endereco',
    'endereco':                    '_endereco',
    'número':                      '_numero',
    'numero':                      '_numero',
    'bairro':                      '_bairro',
    'complemento':                 '_complemento',
    'nome contato':                '_contato_nome',        # ignorado
    'e-mail contato':              '_contato_email',       # ignorado

    # Nomes de campo direto (exportados por este sistema → reimportação)
    'supplier_name':               'supplier_name',
    'supplier_legal_name':         'supplier_legal_name',
    'supplier_legal_identifier':   'supplier_legal_identifier',
    'supplier_email':              'supplier_email',
    'supplier_telephone':          'supplier_telephone',
    'supplier_address':            'supplier_address',
    'description':                 'description',
    'id':                          '_codigo',
}

REQUIRED = {'supplier_name'}   # só o nome é obrigatório; email pode ser gerado


def _norm(s: str) -> str:
    """Remove espaços extras e caixa."""
    return re.sub(r'\s+', ' ', (s or '').strip()).lower()


def _compose_address(raw: dict) -> str | None:
    """Monta endereço completo a partir dos campos separados do template."""
    parts = []
    endereco = raw.get('_endereco', '').strip()
    numero   = raw.get('_numero',   '').strip()
    bairro   = raw.get('_bairro',   '').strip()
    compl    = raw.get('_complemento', '').strip()
    cidade   = raw.get('_cidade',   '').strip()
    estado   = raw.get('_estado',   '').strip()
    cep      = raw.get('_cep',      '').strip()

    if endereco:
        parts.append(endereco + (f', {numero}' if numero else ''))
    if bairro:
        parts.append(bairro)
    if compl:
        parts.append(compl)
    if cidade:
        parts.append(cidade + (f' – {estado}' if estado else ''))
    elif estado:
        parts.append(estado)
    if cep:
        parts.append(f'CEP {cep}')

    return ', '.join(parts) if parts else None


def _clean_cnpj(raw: str) -> str:
    """Remove pontuação extra e mantém só dígitos e separadores CNPJ."""
    val = str(raw or '').strip()
    # Remove sufixos como "/0001-XX" se já formatado com máscara completa
    # Apenas limpa espaços extras
    return val if val else ''


def normalize_row(raw_headers: list, raw_values: list) -> dict | None:
    """
    Recebe os cabeçalhos normalizados → campos internos
    e os valores da linha, retorna o dict pronto para o banco.
    """
    row: dict = {}

    for header, value in zip(raw_headers, raw_values):
        field = HEADER_MAP.get(_norm(header))
        if field:
            val = str(value).strip() if value is not None else ''
            row[field] = val

    # Resolve email: preferir supplier_email; fallback para email_fallback
    if not row.get('supplier_email') and row.get('email_fallback'):
        row['supplier_email'] = row['email_fallback']

    # Monta endereço se não veio direto
    if not row.get('supplier_address'):
        composed = _compose_address(row)
        if composed:
            row['supplier_address'] = composed

    # ID: de '_codigo' (pode ser número ou vazio)
    raw_id = row.get('_codigo', '')
    try:
        row['id'] = int(float(raw_id)) if raw_id else None
    except (ValueError, TypeError):
        row['id'] = None

    # Remove campos internos (prefixo _) antes de retornar
    row = {k: v for k, v in row.items() if not k.startswith('_')}
    # Remove email_fallback
    row.pop('email_fallback', None)

    # Linha sem nome → ignora
    if not row.get('supplier_name', '').strip():
        return None

    return row


def parse_csv(file_path: str) -> list:
    import csv

    rows = []
    with open(file_path, 'r', encoding='utf-8-sig', errors='replace') as f:
        # Detecta separador: ; ou ,
        sample = f.read(2048)
        f.seek(0)
        sep = ';' if sample.count(';') >= sample.count(',') else ','

        reader = csv.reader(f, delimiter=sep)
        headers = [str(h or '') for h in next(reader, [])]

        for raw_values in reader:
            # Preenche colunas faltantes com string vazia
            while len(raw_values) < len(headers):
                raw_values.append('')
            row = normalize_row(headers, raw_values)
            if row:
                rows.append(row)

    return rows


def parse_xlsx(file_path: str) -> list:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)

    # Tenta a aba "Fornecedores"; cai para a primeira
    ws = wb['Fornecedores'] if 'Fornecedores' in wb.sheetnames else wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # Pula linhas de título/subtítulo até encontrar a linha de cabeçalho
    headers = []
    for raw_row in rows_iter:
        candidate = [str(c or '').strip() for c in raw_row]
        # Linha de cabeçalho tem pelo menos 3 células não-vazias
        if sum(1 for c in candidate if c) >= 3:
            headers = candidate
            break

    if not headers:
        return []

    rows = []
    for raw_row in rows_iter:
        raw_values = [str(c or '').strip() for c in raw_row]
        # Linha totalmente vazia → ignora
        if not any(raw_values):
            continue
        row = normalize_row(headers, raw_values)
        if row:
            rows.append(row)

    return rows


def main():
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'Caminho do arquivo não fornecido.'}))
        sys.exit(1)

    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(json.dumps({'error': f'Arquivo não encontrado: {file_path}'}))
        sys.exit(1)

    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == '.csv':
            rows = parse_csv(file_path)
        elif ext in ('.xlsx', '.xls'):
            rows = parse_xlsx(file_path)
        else:
            print(json.dumps({'error': f'Formato não suportado: {ext}. Use .csv ou .xlsx'}))
            sys.exit(1)
    except Exception as e:
        print(json.dumps({'error': f'Erro ao ler arquivo: {str(e)}'}))
        sys.exit(1)

    print(json.dumps({'rows': rows, 'count': len(rows)}, ensure_ascii=False))


if __name__ == '__main__':
    main()
