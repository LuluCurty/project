#!/usr/bin/env python3
"""
export_suppliers.py  <format>
  format: csv | xlsx | pdf

Lê lista de fornecedores do stdin (JSON) e escreve o arquivo em stdout.
Template de colunas segue o padrão importado pelo sistema.
"""
import sys
import json
import io
from datetime import datetime

# ──────────────────────────────────────────────────────────────────────────────
# Colunas que aparecem no template CSV (ordem exata)
# (campo_banco, cabeçalho_csv, largura_xlsx)
# campo_banco=None → calculado / deixado vazio
# ──────────────────────────────────────────────────────────────────────────────
TEMPLATE_COLUMNS = [
    ('supplier_name',             'Nome',                    30),
    ('supplier_legal_name',       'Razão Social',            35),
    ('supplier_legal_identifier', 'CNPJ',                   20),
    ('supplier_email',            'Email',                   30),
    ('description',               'Descrição',              40),
    (None,                        'Status',                   10),   # sempre 'ativado'
    ('id',                        'Código',                  8),
    (None,                        'Inscrição Estadual',      18),
    ('supplier_email',            'Email principal',         30),
    ('supplier_telephone',        'Telefone principal',      20),
    (None,                        'Data fundação/Aniversário', 22),
    (None,                        'CEP',                     12),
    (None,                        'Estado',                  6),
    (None,                        'Cidade',                  20),
    ('supplier_address',          'Endereço',                40),
    (None,                        'Número',                  8),
    (None,                        'Bairro',                  20),
    (None,                        'Complemento',             20),
    (None,                        'Nome Contato',            25),
    (None,                        'E-mail Contato',          30),
]


def _row_values(s: dict) -> list:
    out = []
    for field, _, _ in TEMPLATE_COLUMNS:
        if field is None:
            # Status → 'ativado'; resto fica vazio
            if _ == 'Status':
                out.append('ativado')
            else:
                out.append('')
        else:
            out.append(str(s.get(field) or ''))
    return out


# ──────────────────────────────────────────────────────────────────────────────
# CSV
# ──────────────────────────────────────────────────────────────────────────────
def export_csv(suppliers: list) -> bytes:
    import csv
    out = io.StringIO()
    writer = csv.writer(out, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow([col for _, col, _ in TEMPLATE_COLUMNS])
    for s in suppliers:
        writer.writerow(_row_values(s))
    return ('\ufeff' + out.getvalue()).encode('utf-8')


# ──────────────────────────────────────────────────────────────────────────────
# XLSX
# ──────────────────────────────────────────────────────────────────────────────
def export_xlsx(suppliers: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Fornecedores'

    PRIMARY   = '1E3A5F'
    ALT_ROW   = 'F0F4F8'
    BORDER_C  = 'CCCCCC'

    thin  = Side(style='thin',   color=BORDER_C)
    def make_border(top=thin, bottom=thin, left=thin, right=thin):
        return Border(top=top, bottom=bottom, left=left, right=right)

    hdr_font  = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
    hdr_fill  = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    hdr_border = make_border(top=Side(style='medium', color=PRIMARY))

    cell_font  = Font(size=9, name='Calibri')
    cell_align = Alignment(vertical='center')
    alt_fill   = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type='solid')

    headers = [col for _, col, _ in TEMPLATE_COLUMNS]
    widths  = [w   for _, _, w  in TEMPLATE_COLUMNS]

    # Linha de título
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=len(headers))
    title_cell = ws.cell(row=1, column=1,
                         value=f'Relatório de Fornecedores — {datetime.now().strftime("%d/%m/%Y")}')
    title_cell.font  = Font(bold=True, size=13, color=PRIMARY, name='Calibri')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill  = PatternFill(start_color='EAF0F6', end_color='EAF0F6', fill_type='solid')
    ws.row_dimensions[1].height = 24

    # Linha de subtítulo
    ws.merge_cells(start_row=2, start_column=1, end_row=2,
                   end_column=len(headers))
    sub_cell = ws.cell(row=2, column=1,
                       value=f'{len(suppliers)} fornecedor(es) exportado(s)')
    sub_cell.font      = Font(italic=True, size=9, color='666666', name='Calibri')
    sub_cell.alignment = Alignment(horizontal='center', vertical='center')
    sub_cell.fill      = PatternFill(start_color='EAF0F6', end_color='EAF0F6', fill_type='solid')
    ws.row_dimensions[2].height = 16

    HDR_ROW = 3

    # Cabeçalhos
    for col_i, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=HDR_ROW, column=col_i, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = hdr_border
        ws.column_dimensions[get_column_letter(col_i)].width = width
    ws.row_dimensions[HDR_ROW].height = 26

    # Dados
    for row_i, s in enumerate(suppliers, HDR_ROW + 1):
        fill = alt_fill if row_i % 2 == 0 else None
        for col_i, value in enumerate(_row_values(s), 1):
            cell = ws.cell(row=row_i, column=col_i, value=value)
            cell.font   = cell_font
            cell.alignment = cell_align
            cell.border = make_border()
            if fill:
                cell.fill = fill

    # Freeze panes abaixo do cabeçalho
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    # Auto-filter
    ws.auto_filter.ref = (
        f'A{HDR_ROW}:{get_column_letter(len(headers))}{HDR_ROW + len(suppliers)}'
    )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# PDF  (layout refinado)
# ──────────────────────────────────────────────────────────────────────────────
def export_pdf(suppliers: list) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm, mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, HRFlowable)

    PRIMARY    = colors.HexColor('#1E3A5F')
    SECONDARY  = colors.HexColor('#2E6DA4')
    LIGHT_GREY = colors.HexColor('#F5F5F5')
    BORDER_C   = colors.HexColor('#C8D8E8')
    TEXT_DARK  = colors.HexColor('#1A1A2E')
    TEXT_MUTED = colors.HexColor('#5A6A7A')

    out = io.BytesIO()
    PAGE = landscape(A4)
    L, R, T, B = 1.5*cm, 1.5*cm, 2*cm, 1.5*cm

    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(TEXT_MUTED)
        w, _ = PAGE
        txt = f'Página {doc.page}'
        canvas.drawRightString(w - R, B - 5*mm, txt)
        canvas.drawString(L, B - 5*mm,
                          f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}')
        # linha divisória rodapé
        canvas.setStrokeColor(BORDER_C)
        canvas.setLineWidth(0.5)
        canvas.line(L, B, w - R, B)
        canvas.restoreState()

    doc = SimpleDocTemplate(
        out, pagesize=PAGE,
        leftMargin=L, rightMargin=R, topMargin=T, bottomMargin=B + 8*mm,
        onFirstPage=add_page_number,
        onLaterPages=add_page_number,
    )

    title_style = ParagraphStyle(
        'DocTitle',
        fontName='Helvetica-Bold', fontSize=18,
        textColor=PRIMARY, spaceAfter=2,
        leading=22,
    )
    subtitle_style = ParagraphStyle(
        'DocSub',
        fontName='Helvetica', fontSize=10,
        textColor=TEXT_MUTED, spaceAfter=10,
    )
    section_style = ParagraphStyle(
        'SectionHead',
        fontName='Helvetica-Bold', fontSize=8,
        textColor=colors.white,
    )
    cell_normal = ParagraphStyle(
        'CellNormal',
        fontName='Helvetica', fontSize=7.5,
        textColor=TEXT_DARK, leading=10,
    )
    cell_bold = ParagraphStyle(
        'CellBold',
        fontName='Helvetica-Bold', fontSize=7.5,
        textColor=TEXT_DARK, leading=10,
    )
    cell_muted = ParagraphStyle(
        'CellMuted',
        fontName='Helvetica', fontSize=7,
        textColor=TEXT_MUTED, leading=9,
    )

    # ── Tabela principal ─────────────────────────────────────────────────────
    # Colunas: Nome/Razão Social | CNPJ | Email | Telefone | Endereço | Descrição
    COL_WIDTHS = [5.5*cm, 4*cm, 5*cm, 3.5*cm, 5.5*cm, 4*cm]

    def make_header_para(text):
        return Paragraph(text, section_style)

    header_row = [
        make_header_para('Fornecedor'),
        make_header_para('CNPJ'),
        make_header_para('Email'),
        make_header_para('Telefone'),
        make_header_para('Endereço'),
        make_header_para('Descrição'),
    ]

    data = [header_row]

    for s in suppliers:
        name_para = Paragraph(
            f'<b>{s.get("supplier_name") or "—"}</b>',
            cell_bold,
        )
        legal = s.get('supplier_legal_name') or ''
        if legal:
            name_para = Paragraph(
                f'<b>{s.get("supplier_name") or "—"}</b><br/>'
                f'<font size="6.5" color="#5A6A7A">{legal}</font>',
                cell_normal,
            )

        cnpj  = Paragraph(s.get('supplier_legal_identifier') or '—', cell_normal)
        email = Paragraph(s.get('supplier_email') or '—', cell_normal)
        phone = Paragraph(s.get('supplier_telephone') or '—', cell_normal)
        addr  = Paragraph(s.get('supplier_address')  or '—', cell_muted)
        desc  = Paragraph(s.get('description') or '',        cell_muted)

        data.append([name_para, cnpj, email, phone, addr, desc])

    tbl = Table(data, colWidths=COL_WIDTHS, repeatRows=1)

    row_count = len(data)
    row_backgrounds = []
    for i in range(1, row_count):
        if i % 2 == 0:
            row_backgrounds.append(('BACKGROUND', (0, i), (-1, i), LIGHT_GREY))

    tbl.setStyle(TableStyle([
        # Cabeçalho
        ('BACKGROUND',   (0, 0),  (-1, 0),  PRIMARY),
        ('TEXTCOLOR',    (0, 0),  (-1, 0),  colors.white),
        ('FONTNAME',     (0, 0),  (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0),  (-1, 0),  8),
        ('TOPPADDING',   (0, 0),  (-1, 0),  6),
        ('BOTTOMPADDING',(0, 0),  (-1, 0),  6),
        ('LEFTPADDING',  (0, 0),  (-1, -1), 5),
        ('RIGHTPADDING', (0, 0),  (-1, -1), 5),
        # Dados
        ('FONTSIZE',     (0, 1),  (-1, -1), 7.5),
        ('TOPPADDING',   (0, 1),  (-1, -1), 4),
        ('BOTTOMPADDING',(0, 1),  (-1, -1), 4),
        ('VALIGN',       (0, 0),  (-1, -1), 'MIDDLE'),
        # Grid
        ('LINEBELOW',    (0, 0),  (-1, 0),  1.5, SECONDARY),
        ('INNERGRID',    (0, 1),  (-1, -1), 0.3, BORDER_C),
        ('BOX',          (0, 0),  (-1, -1), 0.5, BORDER_C),
        # Zebra
        *row_backgrounds,
    ]))

    # ── Montar documento ─────────────────────────────────────────────────────
    elements = [
        Paragraph('Relatório de Fornecedores', title_style),
        Paragraph(
            f'{len(suppliers)} fornecedor(es)  •  '
            f'Exportado em {datetime.now().strftime("%d/%m/%Y às %H:%M")}',
            subtitle_style,
        ),
        HRFlowable(width='100%', thickness=2, color=PRIMARY, spaceAfter=10),
        tbl,
    ]

    doc.build(elements,
              onFirstPage=add_page_number,
              onLaterPages=add_page_number)
    return out.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print('Uso: python export_suppliers.py <csv|xlsx|pdf>', file=sys.stderr)
        sys.exit(1)

    fmt       = sys.argv[1].lower()
    suppliers = json.load(sys.stdin)

    handlers = {'csv': export_csv, 'xlsx': export_xlsx, 'pdf': export_pdf}
    if fmt not in handlers:
        print(f'Formato desconhecido: {fmt}', file=sys.stderr)
        sys.exit(1)

    sys.stdout.buffer.write(handlers[fmt](suppliers))


if __name__ == '__main__':
    main()
